import openpyxl
import os

def inspect_excel():
    path = "EE'29 STUDENT LIST.xlsx"
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return
    
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    print(f"Active Sheet Name: {sheet.title}")
    print(f"Max Row: {sheet.max_row}, Max Column: {sheet.max_column}")
    
    # Read first 10 rows
    for r in range(1, 15):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
        print(f"Row {r}: {row_vals}")

if __name__ == "__main__":
    inspect_excel()
