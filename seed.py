import os
import openpyxl
from database import Base, engine, SessionLocal, Student

# Full official student roster for EE'29 batch
DEFAULT_STUDENT_ROSTER = [
    ('2500520200001', 'AARYAN KRISH'),
    ('2500520200002', 'ABHIRAJ CHAURASIA'),
    ('2500520200003', 'ACHHAT SINGH'),
    ('2500520200004', 'ADITYA GUPTA'),
    ('2500520200005', 'ADITYA GUPTA'),
    ('2500520200006', 'ADITYA KUMAR SINGH'),
    ('2500520200007', 'ADITYA TRIPATHI'),
    ('2500520200008', 'AKANKSHA'),
    ('2500520200009', 'AKASH'),
    ('2500520200010', 'AKASH PAL'),
    ('2500520200011', 'AKSHAJ SRIVASTAVA'),
    ('2500520200012', 'ANAND YADAV'),
    ('2500520200013', 'ANIKET KUMAR MISHRA'),
    ('2500520200014', 'ANSHIKA GAUTAM'),
    ('2500520200015', 'AREEB HUSSAIN JAFRI'),
    ('2500520200016', 'ARPIT KUMAR VARSHNEY'),
    ('2500520200017', 'ARUSH SINGH BISEN'),
    ('2500520200018', 'ASHISH KUMAR PANDEY'),
    ('2500520200019', 'DARSHNA RAJPOOT'),
    ('2500520200020', 'DHAWAL SINGH'),
    ('2500520200021', 'DHEERENDRA KUMAR'),
    ('2500520200022', 'DHRUV KUMAR'),
    ('2500520200023', 'DIPESH SINGH'),
    ('2500520200024', 'DIVIJ SAXENA'),
    ('2500520200025', 'DIVYANSHU VERMA'),
    ('2500520200026', 'GAURANSH GUPTA'),
    ('2500520200027', 'GAURAV SINGH'),
    ('2500520200028', 'GYANESHWAR YADAV'),
    ('2500520200029', 'HARSH MITTAL'),
    ('2500520200030', 'HARSH PATHAK'),
    ('2500520200031', 'IPSHITA AWADHIYA'),
    ('2500520200032', 'JATIN VERMA'),
    ('2500520200033', 'JAYDEEP SINGH CHAUHAN'),
    ('2500520200034', 'KARTIK VARSHNEY'),
    ('2500520200035', 'KARTIKEY TYAGI'),
    ('2500520200036', 'KRISHNA CHAURASIYA'),
    ('2500520200037', 'LAVISH YADAV'),
    ('2500520200038', 'LUCKY KUMARI'),
    ('2500520200039', 'MANAK VERMA'),
    ('2500520200040', 'MANASWI GHANGHORIYA'),
    ('2500520200041', 'MD TAHA MIRZA'),
    ('2500520200042', 'NAITIK CHAURASIYA'),
    ('2500520200043', 'OM AGRAWAL'),
    ('2500520200044', 'PIYUSH KUMAR'),
    ('2500520200045', 'PRANJAL'),
    ('2500520200046', 'PRATISHTHA SINGH'),
    ('2500520200047', 'RAJAT UMARVAISHYA'),
    ('2500520200048', 'RIDDHIMA GUPTA'),
    ('2500520200049', 'ROHIT SINGH'),
    ('2500520200050', 'ROHIT VERMA'),
    ('2500520200051', 'SAHIL KUSHWAHA'),
    ('2500520200052', 'SARIM SHAIKH'),
    ('2500520200053', 'SAURABH YADAV'),
    ('2500520200054', 'SHALINI'),
    ('2500520200055', 'SHAURYA SHRESTHA GUPTA'),
    ('2500520200056', 'SHLOK SINGH'),
    ('2500520200057', 'SHOBHI JAISWAL'),
    ('2500520200058', 'SHRESTH RATHOR'),
    ('2500520200059', 'SHREYA SINGH'),
    ('2500520200060', 'SHREYANSH MISHRA'),
    ('2500520200061', 'SHUBHAM MADDHESHIYA'),
    ('2500520200062', 'SHUBHENDU SINGH'),
    ('2500520200063', 'SHUBHRANT SHUKLA'),
    ('2500520200064', 'SIDDHARTH SINGH'),
    ('2500520200065', 'SIDHANT MANI'),
    ('2500520200066', 'SUMIT VERMA'),
    ('2500520200067', 'TANISHKA KATIYAR'),
    ('2500520200068', 'TEJAS'),
    ('2500520200069', 'TUSHAR PACHAURI'),
    ('2500520200070', 'UDIT NARAYAN SINGH'),
    ('2500520200071', 'UPENDRA SINGH'),
    ('2500520200072', 'UTKARSH KUMAR SINGH'),
    ('2500520200073', 'VANDAN DEV BARANWAL'),
    ('2500520200074', 'VARUN PARIHAR'),
    ('2500520200075', 'VED PRAKASH TIWARI'),
    ('2500520200076', 'VISHNU KUMAR'),
    ('2500520200077', 'YASEEN AHMAD SHEIKH')
]

def seed_database(force_reseed: bool = False):
    print("Ensuring database tables exist...")
    Base.metadata.create_all(bind=engine)
    
    db = SessionLocal()
    try:
        current_count = db.query(Student).count()
        excel_file = "EE'29 STUDENT LIST.xlsx"
        
        # Check if database has dummy names ("EE Student 01") or if reseed forced
        has_dummy = db.query(Student).filter(Student.name.like("EE Student %")).first() is not None
        
        if force_reseed or has_dummy or current_count == 0:
            print("Seeding/updating database with official EE'29 student roster...")
            db.query(Student).delete()
            db.commit()
            
            students = []
            if os.path.exists(excel_file):
                print(f"Reading from Excel file '{excel_file}'...")
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                for r in range(2, sheet.max_row + 1):
                    roll_number = sheet.cell(row=r, column=2).value
                    name = sheet.cell(row=r, column=3).value
                    if not roll_number or not name:
                        continue
                    roll_number = str(roll_number).strip()
                    name = str(name).strip()
                    email = f"{roll_number}@ietlucknow.ac.in".lower()
                    students.append(Student(roll_number=roll_number, name=name, email=email))
            else:
                print("Excel file not found. Using embedded official EE'29 student roster...")
                for roll_number, name in DEFAULT_STUDENT_ROSTER:
                    email = f"{roll_number}@ietlucknow.ac.in".lower()
                    students.append(Student(roll_number=roll_number, name=name, email=email))
                    
            db.add_all(students)
            db.commit()
            print(f"Database successfully seeded with {len(students)} official EE'29 student records.")
        else:
            print(f"Database already contains {current_count} valid student records. Skipping re-seeding.")
    except Exception as e:
        db.rollback()
        print(f"An error occurred during seeding: {e}")
    finally:
        db.close()

if __name__ == "__main__":
    seed_database(force_reseed=True)

